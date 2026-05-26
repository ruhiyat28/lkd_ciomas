from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, abort
from flask_login import login_required, current_user
from ..models import db, Nasabah, Pinjaman, Pembayaran, User, hitung_angsuran_bulat, hitung_angsuran_terakhir
from config import Config
from datetime import date, datetime, timedelta
import datetime as dt_module
from dateutil.relativedelta import relativedelta
import io, csv, math, os, zipfile

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

import logging
logger = logging.getLogger(__name__)

import_export_bp = Blueprint('import_export', __name__)

# Whitelist of valid table names for TRUNCATE to prevent SQL injection
TRUNCATE_TABLE_WHITELIST = {
    'jurnal_detail', 'jurnal_umum', 'saldo_awal',
    'pembayaran', 'transaksi_tabungan', 'detail_spp',
    'jaminan_bpkb', 'jaminan_shm', 'jaminan_lain',
    'pinjaman', 'anggota_kelompok', 'rekening_tabungan',
    'nasabah', 'aset',
}

# ─────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────
def _read_excel_rows(file):
    """Read rows from an Excel/CSV file into a list of dicts.

    Returns (rows, error_string) where error_string is None on success.
    Preserves datetime/date objects from openpyxl for parse_date.
    """
    ext = file.filename.rsplit('.', 1)[-1].lower()
    rows = []
    try:
        if ext == 'xlsx':
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                def _cell(v):
                    import datetime as _dt
                    if isinstance(v, (_dt.datetime, _dt.date)):
                        return v
                    return '' if v is None else str(v).strip()
                rows.append(dict(zip(headers, [_cell(v) for v in row])))
        elif ext == 'csv':
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            for row in reader:
                rows.append({k.strip().lower(): v.strip() for k, v in row.items()})
        else:
            return [], 'Format tidak didukung. Gunakan .xlsx atau .csv'
    except Exception as e:
        return [], f'Gagal membaca file: {e}'
    return rows, None
def parse_date(val):
    """Parse tanggal dari berbagai format Excel/CSV secara robust.
    Menangani: datetime object, date object, string ISO, string DD/MM/YYYY,
    string dengan bagian waktu, dan Excel serial number (float/int).
    """
    if val is None or val == '':
        return None

    # 1. Sudah berupa datetime.datetime object (openpyxl SERING return ini untuk cell tanggal)
    if isinstance(val, dt_module.datetime):
        return val.date()

    # 2. Sudah berupa datetime.date object
    if isinstance(val, dt_module.date):
        return val

    # 3. Float — Excel serial number (contoh: 46021.0 = 2026-01-15)
    if isinstance(val, float):
        try:
            n = int(val)
            if 1000 < n < 200000:  # range wajar serial Excel (1902–2147)
                return date(1899, 12, 30) + timedelta(days=n)
        except (ValueError, TypeError):
            pass

    # 4. String — berbagai format
    s = str(val).strip()
    if not s or s.lower() in ('none', 'nan', 'nat', '-', 'null', '0'):
        return None

    # Hapus bagian waktu: "2026-01-15 00:00:00" → "2026-01-15"
    if ' ' in s:
        s = s.split(' ')[0]
    # Hapus timezone: "2026-01-15T00:00:00" → "2026-01-15"
    if 'T' in s:
        s = s.split('T')[0]

    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d',
                '%d/%m/%y', '%m/%d/%Y', '%Y%m%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except (ValueError, TypeError):
            pass

    # Coba Excel serial sebagai string: "46021" atau "46021.0"
    try:
        n = int(float(s))
        if 1000 < n < 200000:
            return date(1899, 12, 30) + timedelta(days=n)
    except (ValueError, TypeError):
        pass

    return None

def parse_int(val, default=0):
    """Parse integer dari berbagai format: Python float/int (dari openpyxl),
    format ribuan Indonesia '1.500.000', format EN '1,500,000', desimal '416700.0'."""
    if val is None or val == '':
        return default
    # Python numeric types langsung dari openpyxl — JANGAN di-str() dulu!
    if isinstance(val, bool):
        return default
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)  # 416700.0 → 416700 ✓
    s = str(val).strip()
    if not s or s.lower() in ('none', 'nan', 'null', '-', ''):
        return default
    try:
        dot_count   = s.count('.')
        comma_count = s.count(',')
        if dot_count == 0 and comma_count == 0:
            return int(s)
        if dot_count > 1 and comma_count == 0:
            # Format ribuan ID: "1.500.000"
            return int(s.replace('.', ''))
        if dot_count == 1 and comma_count == 0:
            # Desimal: "416700.0"
            return int(float(s))
        if dot_count == 0 and comma_count >= 1:
            # Format ribuan EN: "1,500,000"
            return int(s.replace(',', ''))
        # Mixed: "1,500,000.00" (EN dengan desimal)
        return int(float(s.replace(',', '')))
    except (ValueError, TypeError):
        return default

def parse_float(val, default=0.0):
    """Parse float: handle Python float/int, format '1.5' dan '1,5'."""
    if val is None or val == '':
        return default
    if isinstance(val, bool):
        return default
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s.lower() in ('none', 'nan', 'null', '-', ''):
        return default
    try:
        # Ganti koma desimal ke titik (format ID: "1,5" → "1.5")
        # Hati-hati: hanya jika koma adalah desimal, bukan ribuan
        if ',' in s and '.' not in s:
            parts = s.split(',')
            if len(parts) == 2 and len(parts[1]) <= 3:
                s = s.replace(',', '.')
            else:
                s = s.replace(',', '')  # ribuan
        elif ',' in s and '.' in s:
            # "1.500,50" format ID
            s = s.replace('.', '').replace(',', '.')
        return float(s)
    except (ValueError, TypeError):
        return default

def get_next_nasabah_id(kode_desa, start_from=1):
    existing = Nasabah.query.filter_by(kode_desa=kode_desa).all()
    prefix = f"{kode_desa.upper()}-"
    nums = []
    for n in existing:
        try:
            parts = n.nasabah_id.split('-')
            if len(parts) > 1:
                num = int(parts[1])
                if num >= start_from:
                    nums.append(num)
        except (ValueError, TypeError):
            pass
    next_num = max(nums) + 1 if nums else start_from
    return f"{kode_desa.upper()}-{next_num:03d}"

def get_next_spk():
    year  = date.today().year
    prefix = f'SPK-{year}-'
    existing = Pinjaman.query.filter(Pinjaman.spk.like(f'{prefix}%')).all()
    nums = []
    for p in existing:
        try: nums.append(int(p.spk.split('-')[-1]))
        except (ValueError, TypeError): pass
    next_num = max(nums) + 1 if nums else 1
    return f'SPK-{year}-{next_num:05d}'

def get_next_kuitansi():
    year = date.today().strftime('%Y')
    month = date.today().strftime('%m')
    count = Pembayaran.query.filter(Pembayaran.no_kuitansi.like(f'KWT/{year}/{month}/%')).count() + 1
    return f"KWT/{year}/{month}/{count:04d}"

DESA_MAP = {k.upper(): k for k, _ in Config.DESA_LIST}
DESA_NAME_MAP = {v: n for v, n in Config.DESA_LIST}

def resolve_kode_desa(val):
    """Cari kode desa dari kode atau nama"""
    if not val:
        return None, None
    val = str(val).strip().upper()
    # Direct match
    if val in DESA_MAP:
        k = DESA_MAP[val]
        return k, DESA_NAME_MAP[k]
    # Name match
    for kode, nama in Config.DESA_LIST:
        if val in nama.upper() or nama.upper() in val:
            return kode, nama
    return None, None

# ─────────────────────────────────────────
# EXCEL STYLING HELPERS
# ─────────────────────────────────────────
def xl_header_style(ws, row, cols, fill_color="1A56DB"):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF", size=10)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def xl_set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

def _read_xlsx_or_csv(file, required_cols=None):
    """Helper: baca Excel/CSV, return (rows, error_string_or_None)."""
    ext = file.filename.rsplit('.', 1)[-1].lower()
    rows = []
    try:
        if ext == 'xlsx':
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                def _cell(v):
                    import datetime as _dt
                    if isinstance(v, (_dt.datetime, _dt.date)):
                        return v
                    return '' if v is None else str(v).strip()
                rows.append(dict(zip(headers, [_cell(v) for v in row])))
        elif ext == 'csv':
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            for row in reader:
                rows.append({k.strip().lower(): v.strip() for k, v in row.items()})
        else:
            return [], 'Format tidak didukung. Gunakan .xlsx atau .csv'
    except Exception as e:
        return [], f'Gagal membaca file: {e}'
    return rows, None

# ─────────────────────────────────────────
# MAIN PAGE
# ─────────────────────────────────────────
@import_export_bp.route('/')
@login_required
def index():
    stats = {
        'nasabah': Nasabah.query.count(),
        'pinjaman': Pinjaman.query.count(),
        'pembayaran': Pembayaran.query.count(),
    }
    return render_template('import_export/index.html', stats=stats)

# ─────────────────────────────────────────
# DOWNLOAD TEMPLATE
# ─────────────────────────────────────────
@import_export_bp.route('/template/<tipe>')
@login_required
def download_template(tipe):
    if not OPENPYXL_OK:
        flash('Library openpyxl tidak terinstall. Jalankan: pip install openpyxl', 'danger')
        return redirect(url_for('import_export.index'))

    wb = openpyxl.Workbook()
    filename = f"template_import_{tipe}.xlsx"

    if tipe == 'nasabah':
        ws = wb.active
        ws.title = "Nasabah"
        headers = ['nasabah_id','kode_desa','nama','nik','tempat_lahir','tanggal_lahir','jenis_kelamin',
                   'alamat','no_hp','pekerjaan','nama_pasangan','nik_pasangan',
                   'no_hp_pasangan','keterangan_jaminan']
        widths  = [12,10,30,18,18,12,12,35,14,18,25,18,14,30]
        ws.append(headers)
        xl_header_style(ws, 1, len(headers))
        xl_set_widths(ws, widths)
        # Contoh baris
        ws.append(['UT-001','UT','NAMA NASABAH CONTOH','3273010101900001','SERANG',
                   '1990-01-01','Laki-laki','Jl. Contoh No.1 Ciomas','081234567890',
                   'Petani','NAMA PASANGAN','3273010101920002','081234567891',
                   'BPKB Motor Honda Beat 2020'])
        # Petunjuk desa
        ws2 = wb.create_sheet("Kode Desa")
        ws2.append(['Kode','Nama Desa'])
        xl_header_style(ws2, 1, 2)
        for k, n in Config.DESA_LIST:
            ws2.append([k, n])
        filename = "template_import_nasabah.xlsx"

    elif tipe == 'nasabah_kelompok':
        ws = wb.active
        ws.title = "NasabahKelompok"
        headers = ['nasabah_id','kode_desa','nama','no_hp','alamat']
        widths  = [15,10,35,15,40]
        ws.append(headers)
        xl_header_style(ws, 1, len(headers))
        xl_set_widths(ws, widths)
        ws.append(['','UT','NAMA KELOMPOK CONTOH','08123456789','Jl. Contoh No.1'])
        ws2 = wb.create_sheet("Kode Desa")
        ws2.append(['Kode','Nama Desa'])
        xl_header_style(ws2, 1, 2)
        for k, n in Config.DESA_LIST:
            ws2.append([k, n])
        filename = "template_import_nasabah_kelompok.xlsx"
    elif tipe == 'anggota_kelompok':
        ws = wb.active
        ws.title = "AnggotaKelompok"
        headers = ['nasabah_id_kelompok','nama','nik','jabatan','no_hp','alamat']
        widths  = [20,35,18,12,15,40]
        ws.append(headers)
        xl_header_style(ws, 1, len(headers))
        xl_set_widths(ws, widths)
        ws.append(['UT-001','NAMA ANGGOTA','3273010101900001','ketua','08123456789','Jl. Contoh No.1'])
        ws2 = wb.create_sheet("Jabatan")
        ws2.append(['Jabatan','Keterangan'])
        xl_header_style(ws2, 1, 2)
        for j, ket in [('ketua','Ketua Kelompok'),('sekretaris','Sekretaris'),('bendahara','Bendahara'),('anggota','Anggota Biasa')]:
            ws2.append([j, ket])
        filename = "template_import_anggota_kelompok.xlsx"
    elif tipe == 'pinjaman':
        ws = wb.active
        ws.title = "Pinjaman Perorangan"
        headers = ['nasabah_id','spk','jumlah_pinjaman','jasa_persen','tenor',
                   'tujuan','tanggal_pengajuan','tanggal_cair','tanggal_mulai_angsuran',
                   'angsuran_pokok','angsuran_jasa','angsuran_terakhir_pokok','status']
        widths  = [12,20,16,10,8,30,14,14,18,16,12,18,12]
        ws.append(headers)
        xl_header_style(ws, 1, len(headers))
        xl_set_widths(ws, widths)
        today_str   = date.today().strftime('%Y-%m-%d')
        mulai_str   = (date.today() + relativedelta(months=1)).strftime('%Y-%m-%d')
        ws.append(['UT-001','SPK/CONTOH/0001',5000000,1.5,12,
                   'Modal usaha',today_str,today_str,mulai_str,
                   416700,75000,416300,'cair'])
        ws2 = wb.create_sheet("Status Pinjaman")
        ws2.append(['Status','Keterangan'])
        xl_header_style(ws2, 1, 2)
        for s, ket in [('pengajuan','Baru diajukan'),('cek_dokumen','Sedang cek dokumen'),
                       ('verifikasi','Sedang verifikasi lapangan'),('acc_direktur','Menunggu ACC'),
                       ('cair','Sudah dicairkan/aktif'),('lunas','Sudah lunas'),('ditolak','Ditolak')]:
            ws2.append([s, ket])
        ws3 = wb.create_sheet("Petunjuk")
        ws3.append(['Kolom','Keterangan'])
        xl_header_style(ws3, 1, 2)
        for col, ket in [
            ('angsuran_pokok','Angsuran pokok per bulan (isi manual atau kosongkan untuk hitung otomatis)'),
            ('angsuran_jasa','Angsuran jasa per bulan (isi manual atau kosongkan)'),
            ('angsuran_terakhir_pokok','Angsuran pokok bulan terakhir karena pembulatan (kosongkan = hitung otomatis)'),
            ('tanggal_pengajuan','Format: YYYY-MM-DD — boleh kosong (default = hari ini)'),
            ('tanggal_cair','Format: YYYY-MM-DD — tanggal pencairan'),
            ('tanggal_mulai_angsuran','Format: YYYY-MM-DD — biasanya bulan depan setelah cair'),
        ]:
            ws3.append([col, ket])
        filename = "template_import_pinjaman_perorangan.xlsx"

    elif tipe == 'pinjaman_kelompok':
        ws = wb.active
        ws.title = "Pinjaman Kelompok"
        headers = ['nasabah_id_kelompok','spk','jumlah_pinjaman','jasa_persen','tenor',
                   'tujuan','tanggal_pengajuan','tanggal_cair','tanggal_mulai_angsuran',
                   'angsuran_pokok','angsuran_jasa','angsuran_terakhir_pokok','status']
        widths  = [16,22,16,10,8,30,14,14,18,16,12,18,12]
        ws.append(headers)
        xl_header_style(ws, 1, len(headers))
        xl_set_widths(ws, widths)
        today_str = date.today().strftime('%Y-%m-%d')
        mulai_str = (date.today() + relativedelta(months=1)).strftime('%Y-%m-%d')
        # Hitung angsuran contoh: 55jt / 12 bln @ 1.5%
        hasil_kel = hitung_angsuran_bulat(55000000, 12, 1.5)
        ws.append(['CP003-Violet','SPK/CONTOH/0002',55000000,1.5,12,
                   'Modal usaha kelompok',today_str,today_str,mulai_str,
                   hasil_kel['pokok'],hasil_kel['jasa'],hasil_kel['pokok_terakhir'],'cair'])
        # Sheet rincian per anggota
        ws3 = wb.create_sheet("Rincian Anggota")
        ws3.append(['spk','urut','nama_anggota','nik_anggota','jumlah','keterangan'])
        xl_header_style(ws3, 1, 6)
        xl_set_widths(ws3, [22,6,30,20,16,30])
        ws3.append(['SPK/CONTOH/0002',1,'ENTI','3273010101900001',20000000,''])
        ws3.append(['SPK/CONTOH/0002',2,'NURUL FADILAH','3273010101920002',15000000,''])
        ws4 = wb.create_sheet("Status Pinjaman")
        ws4.append(['Status','Keterangan'])
        xl_header_style(ws4, 1, 2)
        for s, ket in [('cair','Sudah dicairkan/aktif'),('lunas','Sudah lunas'),('ditolak','Ditolak')]:
            ws4.append([s, ket])
        ws5 = wb.create_sheet("Petunjuk")
        ws5.append(['Kolom','Keterangan'])
        xl_header_style(ws5, 1, 2)
        for col, ket in [
            ('angsuran_pokok','Angsuran pokok per bulan — kosongkan untuk hitung otomatis'),
            ('angsuran_jasa','Angsuran jasa per bulan — kosongkan untuk hitung otomatis'),
            ('angsuran_terakhir_pokok','Pokok bulan terakhir (karena pembulatan) — kosongkan untuk hitung otomatis'),
            ('tanggal_pengajuan','Format: YYYY-MM-DD — tanggal pengajuan'),
            ('tanggal_cair','Format: YYYY-MM-DD — tanggal pencairan'),
            ('tanggal_mulai_angsuran','Format: YYYY-MM-DD — biasanya sebulan setelah tanggal cair'),
        ]:
            ws5.append([col, ket])
        filename = "template_import_pinjaman_kelompok.xlsx"

    elif tipe == 'pembayaran':
        ws = wb.active
        ws.title = "Pembayaran"
        headers = ['spk','tanggal_bayar','jumlah_bayar','bayar_pokok','bayar_jasa',
                   'angsuran_ke','keterangan']
        widths  = [20,12,14,14,12,10,35]
        ws.append(headers)
        xl_header_style(ws, 1, len(headers))
        xl_set_widths(ws, widths)
        ws.append(['SPK/2024/01/0001','2024-02-20',491700,416700,75000,1,'Angsuran bulan 1'])
        filename = "template_import_pembayaran.xlsx"

    elif tipe == 'tabungan':
        ws = wb.active
        ws.title = "Transaksi Tabungan"
        headers = ['nasabah_id','no_rekening','tanggal','jenis','kategori','jumlah','keterangan']
        widths  = [12,18,12,10,12,14,35]
        ws.append(headers)
        xl_header_style(ws, 1, len(headers))
        xl_set_widths(ws, widths)
        today_str = date.today().strftime('%Y-%m-%d')
        ws.append(['UT-001','',today_str,'setor','sukarela',500000,'Setoran sukarela'])
        ws.append(['UT-001','',today_str,'tarik','sukarela',200000,'Penarikan sukarela'])
        ws2 = wb.create_sheet("Petunjuk")
        ws2.append(['Kolom','Keterangan'])
        xl_header_style(ws2, 1, 2)
        for col, ket in [
            ('nasabah_id','ID nasabah — isi salah satu dengan no_rekening'),
            ('no_rekening','Nomor rekening tabungan — isi salah satu dengan nasabah_id'),
            ('tanggal','Format: YYYY-MM-DD — boleh kosong (default = hari ini)'),
            ('jenis','setor atau tarik'),
            ('kategori','pokok / wajib / sukarela'),
            ('jumlah','Nominal dalam Rupiah (tanpa titik/koma)'),
            ('keterangan','Keterangan transaksi (opsional)'),
        ]:
            ws2.append([col, ket])
        filename = "template_import_tabungan.xlsx"
    else:
        abort(404)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ─────────────────────────────────────────
# IMPORT NASABAH
# ─────────────────────────────────────────
@import_export_bp.route('/import/nasabah', methods=['POST'])
@login_required
def import_nasabah():
    if not current_user.can_write_nasabah():
        abort(403)
    f = request.files.get('file')
    if not f or not f.filename:
        flash('Pilih file terlebih dahulu.', 'danger')
        return redirect(url_for('import_export.index'))

    rows, error = _read_excel_rows(f)
    if error:
        flash(error, 'danger')
        return redirect(url_for('import_export.index'))

    ok, skip, errors = 0, 0, []
    for i, row in enumerate(rows, 2):
        try:
            # ── Validasi minimum: kode_desa + nama ───────────────────
            kode = str(row.get('kode_desa', '')).strip().upper()
            nama = str(row.get('nama', '')).strip().upper()

            if not kode and not nama:
                errors.append(f"Baris {i}: kode_desa dan nama kosong, dilewati.")
                skip += 1
                continue

            if not nama:
                errors.append(f"Baris {i}: kolom nama kosong, dilewati.")
                skip += 1
                continue

            # kode_desa: coba resolve, default ke 'UT' jika tidak dikenal
            kode_desa, nama_desa = resolve_kode_desa(kode)
            if not kode_desa:
                errors.append(f"Baris {i}: kode_desa '{kode}' tidak dikenal, defaultkan ke UT.")
                kode_desa, nama_desa = 'UT', 'UJUNG TEBU'

            nik = str(row.get('nik', '')).strip()

            # NIK duplikat: lewati dengan keterangan
            if nik and Nasabah.query.filter_by(nik=nik).first():
                errors.append(f"Baris {i}: NIK {nik} sudah ada ({nama}), dilewati.")
                skip += 1
                continue

            # Jika NIK kosong, buat NIK sementara unik agar constraint tidak error
            if not nik:
                import uuid
                nik = f"NOID-{uuid.uuid4().hex[:12].upper()}"

            # ID nasabah: manual jika diisi, otomatis jika kosong
            nasabah_id_manual = str(row.get('nasabah_id', '')).strip().upper()
            if nasabah_id_manual and not Nasabah.query.filter_by(nasabah_id=nasabah_id_manual).first():
                nasabah_id = nasabah_id_manual
            else:
                if nasabah_id_manual and Nasabah.query.filter_by(nasabah_id=nasabah_id_manual).first():
                    errors.append(f"Baris {i}: ID {nasabah_id_manual} sudah ada, digenerate otomatis.")
                nasabah_id = get_next_nasabah_id(kode_desa)
            n = Nasabah(
                nasabah_id=nasabah_id,
                kode_desa=kode_desa,
                nama_desa=nama_desa,
                nama=nama,
                jenis='perorangan',
                nik=nik or f"IMPORT-{i}",
                tempat_lahir=str(row.get('tempat_lahir', '')).upper(),
                tanggal_lahir=parse_date(row.get('tanggal_lahir')),
                jenis_kelamin=str(row.get('jenis_kelamin', '')),
                alamat=str(row.get('alamat', '')),
                no_hp=str(row.get('no_hp', '')),
                pekerjaan=str(row.get('pekerjaan', '')),
                nama_pasangan=str(row.get('nama_pasangan', '')).upper(),
                nik_pasangan=str(row.get('nik_pasangan', '')),
                no_hp_pasangan=str(row.get('no_hp_pasangan', '')),
                keterangan_jaminan=str(row.get('keterangan_jaminan', '')),
                created_by=current_user.id,
            )
            db.session.add(n)
            db.session.flush()
            # Auto-buat rekening tabungan
            from ..models import RekeningTabungan as RT
            if not RT.query.filter_by(nasabah_id=n.id).first():
                rek = RT(nasabah_id=n.id, no_rekening=f"TAB-{n.nasabah_id}")
                db.session.add(rek)
            ok += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f"Baris {i}: Error — {e}")
            skip += 1

    db.session.commit()
    msg = f"Import nasabah selesai: {ok} berhasil, {skip} dilewati."
    flash(msg, 'success' if ok > 0 else 'warning')
    if errors:
        flash("Detail: " + " | ".join(errors[:10]), 'warning')
    return redirect(url_for('import_export.index'))

# ─────────────────────────────────────────
# IMPORT PINJAMAN
# ─────────────────────────────────────────
@import_export_bp.route('/import/pinjaman', methods=['POST'])
@login_required
def import_pinjaman():
    if not current_user.can_write_pinjaman():
        abort(403)
    f = request.files.get('file')
    if not f or not f.filename:
        flash('Pilih file terlebih dahulu.', 'danger')
        return redirect(url_for('import_export.index'))

    rows, error = _read_excel_rows(f)
    if error:
        flash(error, 'danger')
        return redirect(url_for('import_export.index'))

    ok, skip, errors = 0, 0, []
    VALID_STATUS = ['pengajuan','cek_dokumen','verifikasi','acc_direktur','cair','lunas','ditolak']

    for i, row in enumerate(rows, 2):
        try:
            # ── Validasi minimum: nasabah_id wajib ada ───────────────
            nasabah_id_str = str(row.get('nasabah_id', '')).strip()
            if not nasabah_id_str:
                errors.append(f"Baris {i}: nasabah_id kosong, dilewati.")
                skip += 1
                continue

            nasabah = Nasabah.query.filter_by(nasabah_id=nasabah_id_str).first()
            if not nasabah:
                errors.append(f"Baris {i}: nasabah_id '{nasabah_id_str}' tidak ditemukan di database, dilewati.")
                skip += 1
                continue

            # SPK: generate otomatis jika kosong, skip jika sudah ada
            spk = str(row.get('spk', '')).strip() or get_next_spk()
            if Pinjaman.query.filter_by(spk=spk).first():
                errors.append(f"Baris {i}: SPK '{spk}' sudah ada, dilewati.")
                skip += 1
                continue

            jumlah      = parse_int(row.get('jumlah_pinjaman', 0))
            jasa_persen = parse_float(row.get('jasa_persen', 1.5))
            tenor       = parse_int(row.get('tenor', 12))
            status      = str(row.get('status', 'cair')).strip().lower()
            if status not in VALID_STATUS:
                status = 'cair'

            # ── Hitung tanggal ────────────────────────────────────────
            tgl_pengajuan = parse_date(row.get('tanggal_pengajuan'))
            tgl_cair      = parse_date(row.get('tanggal_cair'))

            if tgl_cair and not tgl_pengajuan:
                tgl_pengajuan = tgl_cair
            elif tgl_pengajuan and not tgl_cair:
                tgl_cair = tgl_pengajuan
            elif not tgl_pengajuan and not tgl_cair:
                tgl_pengajuan = tgl_cair = date.today()
            else:
                # Jika keduanya diisi, paksa samakan
                tgl_pengajuan = tgl_cair

            tgl_mulai = tgl_cair + relativedelta(months=1)

            # ── Hitung angsuran ───────────────────────────────────────
            angsuran_pokok    = parse_int(row.get('angsuran_pokok', 0))
            angsuran_jasa     = parse_int(row.get('angsuran_jasa', 0))
            angsuran_terakhir = parse_int(row.get('angsuran_terakhir_pokok', 0))

            if jumlah > 0 and tenor > 0:
                if angsuran_pokok > 0:
                    # User memberi angsuran_pokok manual →
                    # hitung jasa otomatis jika kosong, terakhir selalu dihitung ulang
                    if angsuran_jasa == 0:
                        angsuran_jasa = round(jumlah * jasa_persen / 100)
                    # Angsuran terakhir SELALU dihitung ulang agar total = jumlah
                    angsuran_terakhir = hitung_angsuran_terakhir(jumlah, angsuran_pokok, tenor)
                else:
                    # Hitung semua otomatis dengan pembulatan ke atas
                    hasil = hitung_angsuran_bulat(jumlah, tenor, jasa_persen)
                    angsuran_pokok    = hasil['pokok']
                    angsuran_jasa     = hasil['jasa']
                    angsuran_terakhir = hasil['pokok_terakhir']

            p = Pinjaman(
                spk=spk,
                nasabah_id_fk=nasabah.id,
                jumlah_pinjaman=jumlah,
                jasa_persen=jasa_persen,
                tenor=tenor,
                tujuan=str(row.get('tujuan', '')),
                tanggal_pengajuan=tgl_pengajuan,
                tanggal_cair=tgl_cair,
                tanggal_mulai_angsuran=tgl_mulai,
                status=status,
                angsuran_pokok=angsuran_pokok,
                angsuran_jasa=angsuran_jasa,
                angsuran_total=angsuran_pokok + angsuran_jasa,
                angsuran_terakhir_pokok=angsuran_terakhir or angsuran_pokok,
                created_by=current_user.id,
            )
            db.session.add(p)
            db.session.flush()
            ok += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f"Baris {i}: Error — {e}")
            skip += 1

    db.session.commit()
    flash(f"Import pinjaman selesai: {ok} berhasil, {skip} dilewati.", 'success' if ok > 0 else 'warning')
    if errors:
        flash("Detail: " + " | ".join(errors[:10]), 'warning')
    return redirect(url_for('import_export.index'))

# ─────────────────────────────────────────
# IMPORT PINJAMAN KELOMPOK
# ─────────────────────────────────────────
@import_export_bp.route('/import/pinjaman-kelompok', methods=['POST'])
@login_required
def import_pinjaman_kelompok():
    """Import pinjaman kelompok dari Excel (2 sheet: header + rincian anggota)."""
    if not current_user.can_write_pinjaman():
        abort(403)
    f = request.files.get('file')
    if not f or not f.filename:
        flash('Pilih file terlebih dahulu.', 'danger')
        return redirect(url_for('import_export.index'))
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws_main = wb.active
        headers_main = [str(c.value).strip().lower() if c.value else '' for c in next(ws_main.iter_rows(min_row=1, max_row=1))]
        rows_main = []
        for row in ws_main.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            def _c(v):
                import datetime as _dt
                if isinstance(v, (_dt.datetime, _dt.date)): return v
                return '' if v is None else str(v).strip()
            rows_main.append(dict(zip(headers_main, [_c(v) for v in row])))
        # Rincian anggota (sheet kedua, opsional)
        rows_detail = []
        if len(wb.worksheets) > 1:
            ws2 = wb.worksheets[1]
            h2 = [str(c.value).strip().lower() if c.value else '' for c in next(ws2.iter_rows(min_row=1, max_row=1))]
            for row in ws2.iter_rows(min_row=2, values_only=True):
                if not any(row): continue
                rows_detail.append(dict(zip(h2, ['' if v is None else str(v).strip() for v in row])))
    except Exception as e:
        flash(f'Gagal membaca file: {e}', 'danger')
        return redirect(url_for('import_export.index'))

    from ..models import DetailSPP
    VALID_STATUS = ['pengajuan','cek_dokumen','verifikasi','acc_direktur','cair','lunas','ditolak']
    ok, skip, errors = 0, 0, []

    for i, row in enumerate(rows_main, 2):
        try:
            nid_kel = str(row.get('nasabah_id_kelompok', '')).strip().upper()
            if not nid_kel:
                errors.append(f"Baris {i}: nasabah_id_kelompok kosong"); skip += 1; continue
            nasabah = Nasabah.query.filter_by(nasabah_id=nid_kel, jenis='kelompok').first()
            if not nasabah:
                errors.append(f"Baris {i}: kelompok '{nid_kel}' tidak ditemukan"); skip += 1; continue
            spk = str(row.get('spk', '')).strip() or get_next_spk()
            if Pinjaman.query.filter_by(spk=spk).first():
                errors.append(f"Baris {i}: SPK '{spk}' sudah ada"); skip += 1; continue
            jumlah      = parse_int(row.get('jumlah_pinjaman', 0))
            jasa_persen = parse_float(row.get('jasa_persen', 1.5))
            tenor       = parse_int(row.get('tenor', 12))
            status      = str(row.get('status', 'cair')).strip().lower()
            if status not in VALID_STATUS: status = 'cair'

            # ── Hitung tanggal ────────────────────────────────────────
            tgl_pengajuan = parse_date(row.get('tanggal_pengajuan'))
            tgl_cair      = parse_date(row.get('tanggal_cair'))

            if tgl_cair and not tgl_pengajuan:
                tgl_pengajuan = tgl_cair
            elif tgl_pengajuan and not tgl_cair:
                tgl_cair = tgl_pengajuan
            elif not tgl_pengajuan and not tgl_cair:
                tgl_pengajuan = tgl_cair = date.today()
            else:
                tgl_pengajuan = tgl_cair

            tgl_mulai = tgl_cair + relativedelta(months=1)

            # ── Hitung angsuran (sama logic dengan perorangan) ────────────
            angsuran_pokok    = parse_int(row.get('angsuran_pokok', 0))
            angsuran_jasa     = parse_int(row.get('angsuran_jasa', 0))
            angsuran_terakhir = parse_int(row.get('angsuran_terakhir_pokok', 0))

            if jumlah > 0 and tenor > 0:
                if angsuran_pokok > 0:
                    if angsuran_jasa == 0:
                        angsuran_jasa = round(jumlah * jasa_persen / 100)
                    angsuran_terakhir = hitung_angsuran_terakhir(jumlah, angsuran_pokok, tenor)
                else:
                    hasil = hitung_angsuran_bulat(jumlah, tenor, jasa_persen)
                    angsuran_pokok    = hasil['pokok']
                    angsuran_jasa     = hasil['jasa']
                    angsuran_terakhir = hasil['pokok_terakhir']

            p = Pinjaman(
                spk=spk, nasabah_id_fk=nasabah.id,

                jumlah_pinjaman=jumlah, jasa_persen=jasa_persen, tenor=tenor,
                tujuan=str(row.get('tujuan', '')),
                tanggal_pengajuan=tgl_pengajuan,
                tanggal_cair=tgl_cair,
                tanggal_mulai_angsuran=tgl_mulai,
                status=status,
                angsuran_pokok=angsuran_pokok, angsuran_jasa=angsuran_jasa,
                angsuran_total=angsuran_pokok + angsuran_jasa,
                angsuran_terakhir_pokok=angsuran_terakhir or angsuran_pokok,
                jenis_pinjaman='spp', created_by=current_user.id,
            )
            db.session.add(p); db.session.flush()
            # Masukkan rincian anggota jika ada
            anggota_spk = [r for r in rows_detail if str(r.get('spk','')).strip().upper() == spk.upper()]
            for j, det in enumerate(anggota_spk, 1):
                db.session.add(DetailSPP(
                    pinjaman_id=p.id, urut=parse_int(det.get('urut', j)) or j,
                    nama_anggota=str(det.get('nama_anggota','')).strip().upper(),
                    nik_anggota=str(det.get('nik_anggota','')).strip(),
                    jumlah=parse_int(det.get('jumlah', 0)),
                    keterangan=str(det.get('keterangan','')).strip(),
                ))
            ok += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f"Baris {i}: {e}")
            skip += 1

    db.session.commit()
    flash(f"Import pinjaman kelompok: {ok} berhasil, {skip} dilewati.", 'success' if ok > 0 else 'warning')
    if errors: flash("Detail: " + " | ".join(errors[:10]), 'warning')
    return redirect(url_for('import_export.index'))

# ─────────────────────────────────────────
# BERSIHKAN DATA
# ─────────────────────────────────────────
@import_export_bp.route('/bersihkan-data', methods=['GET','POST'])
@login_required
def bersihkan_data():
    if current_user.role != 'admin':
        abort(403)
    from ..models import (Pembayaran, Pinjaman, Nasabah, AnggotaKelompok,
                          DetailSPP, JurnalUmum, JurnalDetail, Aset)
    from sqlalchemy import text
    if request.method == 'POST':
        password_konfirmasi = request.form.get('password', '')
        if not current_user.check_password(password_konfirmasi):
            flash('Password salah. Data tidak dihapus.', 'danger')
            return redirect(url_for('import_export.bersihkan_data'))
        modul = request.form.getlist('modul[]')
        deleted = {}
        try:
            if db.engine.name == 'postgresql':
                truncate_tables = []
                if 'jurnal' in modul: truncate_tables.extend(['jurnal_detail', 'jurnal_umum', 'saldo_awal'])
                if 'pembayaran' in modul: truncate_tables.append('pembayaran')
                if 'pinjaman' in modul: truncate_tables.extend(['detail_spp', 'jaminan_bpkb', 'jaminan_shm', 'jaminan_lain', 'pinjaman'])
                if 'nasabah' in modul: truncate_tables.extend(['anggota_kelompok', 'rekening_tabungan', 'transaksi_tabungan', 'nasabah'])
                if 'aset' in modul: truncate_tables.append('aset')

                # Validate table names against whitelist to prevent SQL injection
                validated_tables = []
                for t in truncate_tables:
                    if t not in TRUNCATE_TABLE_WHITELIST:
                        logger.warning('Blocked attempt to truncate unlisted table: %s', t)
                        continue
                    validated_tables.append(t)

                if validated_tables:
                    sql = f"TRUNCATE TABLE {', '.join(validated_tables)} CASCADE"
                    db.session.execute(text(sql))
                    deleted['Status'] = 'Dibersihkan via TRUNCATE'
            else:
                # Fallback non-Postgres (SQLite dll)
                if 'pembayaran' in modul:
                    deleted['Pembayaran'] = Pembayaran.query.count()
                    from ..models import TransaksiTabungan
                    TransaksiTabungan.query.delete()
                    Pembayaran.query.delete()
                if 'jurnal' in modul:
                    deleted['Jurnal Detail'] = JurnalDetail.query.count()
                    deleted['Jurnal Umum'] = JurnalUmum.query.count()
                    JurnalDetail.query.delete()
                    JurnalUmum.query.delete()
                    db.session.execute(text("DELETE FROM saldo_awal"))
                if 'pinjaman' in modul:
                    from ..models import JaminanBPKB, JaminanSHM, JaminanLain
                    JaminanBPKB.query.delete()
                    JaminanSHM.query.delete()
                    JaminanLain.query.delete()
                    deleted['Detail SPP'] = DetailSPP.query.count()
                    deleted['Pinjaman'] = Pinjaman.query.count()
                    DetailSPP.query.delete()
                    Pinjaman.query.delete()
                if 'nasabah' in modul:
                    from ..models import RekeningTabungan, TransaksiTabungan
                    deleted['Anggota'] = AnggotaKelompok.query.count()
                    deleted['Nasabah'] = Nasabah.query.count()
                    AnggotaKelompok.query.delete()
                    # Rekening tabungan dan transaksi harus dihapus jika nasabah dihapus
                    TransaksiTabungan.query.delete()
                    RekeningTabungan.query.delete()
                    Nasabah.query.delete()
                if 'aset' in modul:
                    deleted['Aset'] = Aset.query.count()
                    Aset.query.delete()
            db.session.commit()
            
            # ── SAFETY: Pastikan admin tetap ada setelah pembersihan ──
            admin = User.query.filter_by(username='admin').first()
            if not admin:
                admin = User(username='admin', nama_lengkap='Administrator', role='admin')
                db.session.add(admin)
            admin.set_password('admin123')
            admin.aktif = True
            db.session.commit()
            # ────────────────────────────────────────────────────────

            detail = ', '.join([f"{v} {k}" for k,v in deleted.items()])
            flash(f'Data berhasil dibersihkan: {detail}. Password admin telah diatur ulang ke "admin123".', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal: {e}', 'danger')
        return redirect(url_for('import_export.index'))
    return render_template('import_export/bersihkan_data.html')

# ─────────────────────────────────────────
# IMPORT PEMBAYARAN
# ─────────────────────────────────────────
@import_export_bp.route('/import/pembayaran', methods=['POST'])
@login_required
def import_pembayaran():
    if current_user.role not in ['admin', 'kasir']:
        abort(403)
    f = request.files.get('file')
    if not f or not f.filename:
        flash('Pilih file terlebih dahulu.', 'danger')
        return redirect(url_for('import_export.index'))

    rows, error = _read_excel_rows(f)
    if error:
        flash(error, 'danger')
        return redirect(url_for('import_export.index'))

    ok, skip, errors = 0, 0, []
    for i, row in enumerate(rows, 2):
        try:
            spk = str(row.get('spk', '')).strip()
            pinjaman = Pinjaman.query.filter_by(spk=spk).first()
            if not pinjaman:
                errors.append(f"Baris {i}: SPK '{spk}' tidak ditemukan, dilewati.")
                skip += 1
                continue

            jumlah_bayar = parse_int(row.get('jumlah_bayar', 0))
            # Permisif: jika jumlah_bayar 0 tapi ada alokasi pokok/jasa, hitung total
            bayar_pokok_tmp = parse_int(row.get('bayar_pokok', 0))
            bayar_jasa_tmp  = parse_int(row.get('bayar_jasa', 0))
            if jumlah_bayar <= 0:
                jumlah_bayar = bayar_pokok_tmp + bayar_jasa_tmp
            if jumlah_bayar <= 0:
                errors.append(f"Baris {i}: jumlah_bayar tidak valid (0), dilewati.")
                skip += 1
                continue

            no_kuitansi = get_next_kuitansi()
            bayar_pokok = parse_int(row.get('bayar_pokok', 0))
            bayar_jasa  = parse_int(row.get('bayar_jasa', 0))

            # Jika pokok+jasa tidak diisi, estimasi otomatis
            if bayar_pokok == 0 and bayar_jasa == 0:
                bayar_jasa  = min(jumlah_bayar, pinjaman.angsuran_jasa or 0)
                bayar_pokok = max(0, jumlah_bayar - bayar_jasa)

            pb = Pembayaran(
                no_kuitansi=no_kuitansi,
                pinjaman_id=pinjaman.id,
                tanggal_bayar=parse_date(row.get('tanggal_bayar')) or date.today(),
                jumlah_bayar=jumlah_bayar,
                bayar_pokok=bayar_pokok,
                bayar_jasa=bayar_jasa,
                angsuran_ke=parse_int(row.get('angsuran_ke', 0)) or None,
                keterangan=str(row.get('keterangan', 'Import data')),
                created_by=current_user.id,
            )
            db.session.add(pb)

            # Update status pinjaman jika lunas
            total_pokok = sum(p2.bayar_pokok for p2 in pinjaman.pembayaran) + bayar_pokok
            if total_pokok >= pinjaman.jumlah_pinjaman:
                pinjaman.status = 'lunas'

            db.session.flush()
            ok += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f"Baris {i}: Error — {e}")
            skip += 1

    db.session.commit()
    flash(f"Import pembayaran selesai: {ok} berhasil, {skip} dilewati.", 'success' if ok > 0 else 'warning')
    if errors:
        flash("Detail: " + " | ".join(errors[:10]), 'warning')
    return redirect(url_for('import_export.index'))

# ─────────────────────────────────────────
# IMPORT TRANSAKSI TABUNGAN
# ─────────────────────────────────────────
@import_export_bp.route('/import/tabungan', methods=['POST'])
@login_required
def import_tabungan():
    """Import transaksi tabungan dari Excel/CSV.
    Kolom: nasabah_id/no_rekening*, tanggal*, jenis(setor/tarik)*, kategori(pokok/wajib/sukarela)*, jumlah*, keterangan
    """
    if not current_user.is_admin() and not current_user.is_keuangan() and not current_user.is_kasir():
        abort(403)
    f = request.files.get('file')
    if not f or not f.filename:
        flash('Pilih file terlebih dahulu.', 'danger')
        return redirect(url_for('import_export.index'))

    from ..models import RekeningTabungan, TransaksiTabungan

    rows, error = _read_excel_rows(f)
    if error:
        flash(error, 'danger')
        return redirect(url_for('import_export.index'))

    ok, skip, errors = 0, 0, []
    JENIS_VALID = ['setor', 'tarik']
    KATEGORI_VALID = ['pokok', 'wajib', 'sukarela']

    for i, row in enumerate(rows, 2):
        try:
            # Resolve rekening
            nasabah_id_str = str(row.get('nasabah_id', '')).strip()
            no_rek_str = str(row.get('no_rekening', '')).strip()

            rek = None
            if no_rek_str:
                rek = RekeningTabungan.query.filter_by(no_rekening=no_rek_str).first()
            elif nasabah_id_str:
                nasabah = Nasabah.query.filter_by(nasabah_id=nasabah_id_str).first()
                if nasabah:
                    rek = RekeningTabungan.query.filter_by(nasabah_id=nasabah.id).first()

            if not rek:
                errors.append(f"Baris {i}: Rekening tidak ditemukan (nasabah_id='{nasabah_id_str}', no_rek='{no_rek_str}')")
                skip += 1
                continue

            # Validasi jenis & kategori
            jenis = str(row.get('jenis', '')).strip().lower()
            if jenis not in JENIS_VALID:
                errors.append(f"Baris {i}: jenis '{jenis}' tidak valid (harus setor/tarik)")
                skip += 1
                continue

            kategori = str(row.get('kategori', '')).strip().lower()
            if kategori not in KATEGORI_VALID:
                errors.append(f"Baris {i}: kategori '{kategori}' tidak valid (harus pokok/wajib/sukarela)")
                skip += 1
                continue

            jumlah = parse_int(row.get('jumlah', 0))
            if jumlah <= 0:
                errors.append(f"Baris {i}: jumlah tidak valid (0 atau negatif)")
                skip += 1
                continue

            # Cek saldo untuk penarikan
            if jenis == 'tarik':
                saldo_field = f'saldo_{kategori}'
                saldo_aktif = getattr(rek, saldo_field, 0)
                if saldo_aktif < jumlah:
                    errors.append(f"Baris {i}: Saldo {kategori} tidak cukup (ada {saldo_aktif}, tarik {jumlah})")
                    skip += 1
                    continue

            tgl = parse_date(row.get('tanggal')) or date.today()
            keterangan = str(row.get('keterangan', '')).strip() or 'Import data'

            # Buat no bukti otomatis
            year = tgl.year
            month = tgl.month
            count = TransaksiTabungan.query.filter(
                TransaksiTabungan.tanggal >= date(year, month, 1),
                TransaksiTabungan.tanggal < (date(year, month, 1) + relativedelta(months=1))
            ).count() + 1
            no_bukti = f"TAB/{year}/{month:02d}/{count:04d}"

            # Buat transaksi
            trx = TransaksiTabungan(
                rekening_id=rek.id,
                tanggal=tgl,
                jenis=jenis,
                kategori=kategori,
                jumlah=jumlah,
                keterangan=keterangan,
                no_bukti=no_bukti,
                created_by=current_user.id,
            )
            db.session.add(trx)

            # Update saldo
            saldo_field = f'saldo_{kategori}'
            if jenis == 'setor':
                setattr(rek, saldo_field, getattr(rek, saldo_field, 0) + jumlah)
            else:
                setattr(rek, saldo_field, getattr(rek, saldo_field, 0) - jumlah)

            ok += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f"Baris {i}: Error — {e}")
            skip += 1

    db.session.commit()
    flash(f"Import transaksi tabungan selesai: {ok} berhasil, {skip} dilewati.", 'success' if ok > 0 else 'warning')
    if errors:
        flash("Detail: " + " | ".join(errors[:10]), 'warning')
    return redirect(url_for('import_export.index'))


# ─────────────────────────────────────────
# EXPORT NASABAH
# ─────────────────────────────────────────
@import_export_bp.route('/export/nasabah')
@login_required
def export_nasabah():
    fmt = request.args.get('fmt', 'xlsx')
    desa = request.args.get('desa', '')
    q = Nasabah.query
    if desa:
        q = q.filter_by(kode_desa=desa)
    data = q.order_by(Nasabah.nasabah_id).all()

    headers = ['ID Nasabah','Kode Desa','Nama Desa','Nama','NIK','Tempat Lahir',
               'Tanggal Lahir','Jenis Kelamin','Alamat','No. HP','Pekerjaan',
               'Nama Pasangan','NIK Pasangan','No. HP Pasangan','Keterangan Jaminan',
               'Foto','KTP','KK','SKU','Penghasilan','Jaminan','Tgl Daftar']

    rows = []
    for n in data:
        rows.append([
            n.nasabah_id, n.kode_desa, n.nama_desa, n.nama, n.nik,
            n.tempat_lahir or '', n.tanggal_lahir.strftime('%Y-%m-%d') if n.tanggal_lahir else '',
            n.jenis_kelamin or '', n.alamat or '', n.no_hp or '', n.pekerjaan or '',
            n.nama_pasangan or '', n.nik_pasangan or '', n.no_hp_pasangan or '',
            n.keterangan_jaminan or '',
            'Ada' if n.foto else '', 'Ada' if n.ktp else '',
            'Ada' if n.kk else '', 'Ada' if n.surat_usaha else '',
            'Ada' if n.bukti_penghasilan else '', 'Ada' if n.jaminan else '',
            n.created_at.strftime('%Y-%m-%d') if n.created_at else '',
        ])

    return _send_export(headers, rows, fmt, f"export_nasabah_{date.today()}")

# ─────────────────────────────────────────
# EXPORT PINJAMAN
# ─────────────────────────────────────────
@import_export_bp.route('/export/pinjaman')
@login_required
def export_pinjaman():
    fmt = request.args.get('fmt', 'xlsx')
    status = request.args.get('status', '')
    desa   = request.args.get('desa', '')
    q = Pinjaman.query.join(Nasabah)
    if status:
        q = q.filter(Pinjaman.status == status)
    if desa:
        q = q.filter(Nasabah.kode_desa == desa)
    data = q.order_by(Pinjaman.id).all()

    headers = ['No. SPK','ID Nasabah','Nama','Desa','Jumlah Pinjaman','Jasa %/Bln',
               'Tenor','Tgl Pengajuan','Tgl Cair','Tgl Mulai Angsuran',
               'Angsuran Pokok','Angsuran Jasa','Angsuran Total','Angsuran Terakhir Pokok',
               'Status','Terbayar Pokok','Terbayar Jasa','Saldo Pokok',
               'Kolektibilitas','Bulan Tunggak','Tujuan']

    rows = []
    for p in data:
        tp, tj = p.get_realisasi_pembayaran()
        kolek, kolek_lbl = p.get_kolektibilitas()
        _, _, bn = p.get_tunggakan()
        rows.append([
            p.spk, p.nasabah.nasabah_id, p.nasabah.nama, p.nasabah.nama_desa,
            p.jumlah_pinjaman, p.jasa_persen, p.tenor,
            p.tanggal_pengajuan.strftime('%Y-%m-%d') if p.tanggal_pengajuan else '',
            p.tanggal_cair.strftime('%Y-%m-%d') if p.tanggal_cair else '',
            p.tanggal_mulai_angsuran.strftime('%Y-%m-%d') if p.tanggal_mulai_angsuran else '',
            p.angsuran_pokok or 0, p.angsuran_jasa or 0, p.angsuran_total or 0,
            p.angsuran_terakhir_pokok or 0,
            p.status, tp, tj, p.get_saldo_pokok(),
            f"Kol {kolek}: {kolek_lbl}", bn, p.tujuan or '',
        ])

    return _send_export(headers, rows, fmt, f"export_pinjaman_{date.today()}")

# ─────────────────────────────────────────
# EXPORT PEMBAYARAN
# ─────────────────────────────────────────
@import_export_bp.route('/export/pembayaran')
@login_required
def export_pembayaran():
    fmt     = request.args.get('fmt', 'xlsx')
    dari    = request.args.get('dari', '')
    sampai  = request.args.get('sampai', '')
    desa    = request.args.get('desa', '')
    q = Pembayaran.query.join(Pinjaman).join(Nasabah)
    if dari:
        try:
            q = q.filter(Pembayaran.tanggal_bayar >= datetime.strptime(dari, '%Y-%m-%d').date())
        except (ValueError, TypeError):
            pass
    if sampai:
        try:
            q = q.filter(Pembayaran.tanggal_bayar <= datetime.strptime(sampai, '%Y-%m-%d').date())
        except (ValueError, TypeError):
            pass
    if desa:
        q = q.filter(Nasabah.kode_desa == desa)
    data = q.order_by(Pembayaran.tanggal_bayar).all()

    headers = ['No. Kuitansi','Tgl Bayar','ID Nasabah','Nama','Desa','No. SPK',
               'Jumlah Bayar','Bayar Pokok','Bayar Jasa','Angsuran Ke','Keterangan']
    rows = []
    for pb in data:
        rows.append([
            pb.no_kuitansi,
            pb.tanggal_bayar.strftime('%Y-%m-%d'),
            pb.pinjaman.nasabah.nasabah_id,
            pb.pinjaman.nasabah.nama,
            pb.pinjaman.nasabah.nama_desa,
            pb.pinjaman.spk,
            pb.jumlah_bayar, pb.bayar_pokok, pb.bayar_jasa,
            pb.angsuran_ke or '', pb.keterangan or '',
        ])

    return _send_export(headers, rows, fmt, f"export_pembayaran_{date.today()}")

# ─────────────────────────────────────────
# EXPORT KOLEKTIBILITAS
# ─────────────────────────────────────────
@import_export_bp.route('/export/kolektibilitas')
@login_required
def export_kolektibilitas():
    fmt  = request.args.get('fmt', 'xlsx')
    desa = request.args.get('desa', '')
    q = Pinjaman.query.join(Nasabah).filter(Pinjaman.status == 'cair')
    if desa:
        q = q.filter(Nasabah.kode_desa == desa)
    data = q.all()

    headers = ['No. SPK','ID Nasabah','Nama','Desa','Jumlah Pinjaman','Saldo Pokok',
               'Bulan Tunggak','Kolektibilitas','Cadangan Risiko %','Cadangan Risiko (Rp)']
    rows = []
    for p in data:
        kolek, lbl = p.get_kolektibilitas()
        saldo = p.get_saldo_pokok()
        cad_pct = Config.KOLEK_CADANGAN[kolek]
        _, _, bn = p.get_tunggakan()
        rows.append([
            p.spk, p.nasabah.nasabah_id, p.nasabah.nama, p.nasabah.nama_desa,
            p.jumlah_pinjaman, saldo, bn,
            f"Kol {kolek}: {lbl}", cad_pct * 100, round(saldo * cad_pct),
        ])
    rows.sort(key=lambda x: x[6], reverse=True)

    return _send_export(headers, rows, fmt, f"export_kolektibilitas_{date.today()}")

# ─────────────────────────────────────────
# EXPORT TABUNGAN
# ─────────────────────────────────────────
@import_export_bp.route('/export/tabungan')
@login_required
def export_tabungan():
    fmt = request.args.get('fmt', 'xlsx')
    from ..models import RekeningTabungan
    data = RekeningTabungan.query.join(Nasabah).order_by(RekeningTabungan.no_rekening).all()

    headers = ['No. Rekening','ID Nasabah','Nama','Desa','Saldo Pokok','Saldo Wajib','Saldo Sukarela','Total Saldo']
    rows = []
    for r in data:
        rows.append([
            r.no_rekening, r.nasabah.nasabah_id, r.nasabah.nama, r.nasabah.nama_desa,
            r.saldo_pokok, r.saldo_wajib, r.saldo_sukarela, r.total_saldo()
        ])

    return _send_export(headers, rows, fmt, f"export_tabungan_{date.today()}")

# ─────────────────────────────────────────
# EXPORT ZIP (semua sekaligus)
# ─────────────────────────────────────────
@import_export_bp.route('/export/semua')
@login_required
def export_semua():
    fmt = request.args.get('fmt', 'xlsx')
    buf_zip = io.BytesIO()
    with zipfile.ZipFile(buf_zip, 'w', zipfile.ZIP_DEFLATED) as zf:
        for nama, endpoint in [
            ('nasabah', export_nasabah),
            ('pinjaman', export_pinjaman),
            ('pembayaran', export_pembayaran),
            ('kolektibilitas', export_kolektibilitas),
            ('tabungan', export_tabungan),
        ]:
            # Kita panggil helper langsung
            pass

    # Simpler: generate each buffer individually
    buf_zip = io.BytesIO()
    ext = 'xlsx' if fmt == 'xlsx' else 'csv'
    with zipfile.ZipFile(buf_zip, 'w', zipfile.ZIP_DEFLATED) as zf:
        for label, headers, rows in _get_all_export_data():
            fbuf = io.BytesIO()
            if fmt == 'xlsx' and OPENPYXL_OK:
                _write_xlsx(fbuf, headers, rows)
            else:
                _write_csv(fbuf, headers, rows)
            fbuf.seek(0)
            zf.writestr(f"{label}.{ext}", fbuf.read())

    buf_zip.seek(0)
    return send_file(buf_zip, as_attachment=True,
                     download_name=f"export_lkd_ciomas_{date.today()}.zip",
                     mimetype='application/zip')

def _get_all_export_data():
    # Nasabah
    hdrs = ['ID Nasabah','Kode Desa','Nama Desa','Nama','NIK','Tempat Lahir','Tanggal Lahir',
            'Jenis Kelamin','Alamat','No. HP','Pekerjaan','Nama Pasangan','NIK Pasangan',
            'No. HP Pasangan','Keterangan Jaminan']
    rows = []
    for n in Nasabah.query.order_by(Nasabah.nasabah_id).all():
        rows.append([n.nasabah_id,n.kode_desa,n.nama_desa,n.nama,n.nik,
                     n.tempat_lahir or '',n.tanggal_lahir.strftime('%Y-%m-%d') if n.tanggal_lahir else '',
                     n.jenis_kelamin or '',n.alamat or '',n.no_hp or '',n.pekerjaan or '',
                     n.nama_pasangan or '',n.nik_pasangan or '',n.no_hp_pasangan or '',n.keterangan_jaminan or ''])
    yield 'nasabah', hdrs, rows

    # Pinjaman
    hdrs = ['No. SPK','ID Nasabah','Nama','Desa','Jumlah Pinjaman','Jasa %','Tenor',
            'Tgl Cair','Status','Saldo Pokok','Kolektibilitas']
    rows = []
    for p in Pinjaman.query.join(Nasabah).order_by(Pinjaman.id).all():
        kolek, lbl = p.get_kolektibilitas()
        rows.append([p.spk,p.nasabah.nasabah_id,p.nasabah.nama,p.nasabah.nama_desa,
                     p.jumlah_pinjaman,p.jasa_persen,p.tenor,
                     p.tanggal_cair.strftime('%Y-%m-%d') if p.tanggal_cair else '',
                     p.status,p.get_saldo_pokok(),f"Kol {kolek}: {lbl}"])
    yield 'pinjaman', hdrs, rows

    # Pembayaran
    hdrs = ['No. Kuitansi','Tgl Bayar','Nama','Desa','No. SPK','Jumlah Bayar','Pokok','Jasa']
    rows = []
    for pb in Pembayaran.query.join(Pinjaman).join(Nasabah).order_by(Pembayaran.tanggal_bayar).all():
        rows.append([pb.no_kuitansi,pb.tanggal_bayar.strftime('%Y-%m-%d'),
                     pb.pinjaman.nasabah.nama,pb.pinjaman.nasabah.nama_desa,
                     pb.pinjaman.spk,pb.jumlah_bayar,pb.bayar_pokok,pb.bayar_jasa])
    yield 'pembayaran', hdrs, rows

    # Tabungan
    from ..models import RekeningTabungan
    hdrs = ['No. Rekening','ID Nasabah','Nama','Desa','Saldo Pokok','Saldo Wajib','Saldo Sukarela','Total Saldo']
    rows = []
    for r in RekeningTabungan.query.join(Nasabah).order_by(RekeningTabungan.no_rekening).all():
        rows.append([r.no_rekening, r.nasabah.nasabah_id, r.nasabah.nama, r.nasabah.nama_desa,
                     r.saldo_pokok, r.saldo_wajib, r.saldo_sukarela, r.total_saldo()])
    yield 'tabungan', hdrs, rows

# ─────────────────────────────────────────
# HELPERS: write & send
# ─────────────────────────────────────────
def _write_xlsx(buf, headers, rows):
    if not OPENPYXL_OK:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    xl_header_style(ws, 1, len(headers))
    for row in rows:
        ws.append(row)
    # Auto-width approx
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    wb.save(buf)

def _write_csv(buf, headers, rows):
    text = io.TextIOWrapper(buf, encoding='utf-8-sig', newline='')
    writer = csv.writer(text)
    writer.writerow(headers)
    writer.writerows(rows)
    text.detach()

def _send_export(headers, rows, fmt, basename):
    buf = io.BytesIO()
    if fmt == 'csv':
        _write_csv(buf, headers, rows)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name=f"{basename}.csv",
                         mimetype='text/csv; charset=utf-8')
    else:
        if not OPENPYXL_OK:
            flash('openpyxl tidak terinstall. Gunakan format CSV.', 'warning')
            _write_csv(buf, headers, rows)
            buf.seek(0)
            return send_file(buf, as_attachment=True, download_name=f"{basename}.csv",
                             mimetype='text/csv')
        _write_xlsx(buf, headers, rows)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name=f"{basename}.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@import_export_bp.route('/import/nasabah-kelompok', methods=['POST'])
@login_required
def import_nasabah_kelompok():
    """Import nasabah kelompok dari Excel."""
    if not current_user.can_write_nasabah(): abort(403)
    from ..models import Nasabah, RekeningTabungan
    file = request.files.get('file')
    if not file or not file.filename:
        flash('Pilih file terlebih dahulu.','warning')
        return redirect(url_for('import_export.index'))

    rows, err = _read_xlsx_or_csv(file, ['kode_desa','nama','no_hp','alamat','nasabah_id'])
    if err:
        flash(err,'danger'); return redirect(url_for('import_export.index'))

    ok = 0; errors = []
    for i, row in enumerate(rows,2):
        nama = str(row.get('nama','')).strip().upper()
        kode_desa, nama_desa = resolve_kode_desa(str(row.get('kode_desa','')).strip().upper())
        if not nama or not kode_desa:
            errors.append(f"Baris {i}: nama/desa kosong"); continue
        nama_desa = nama_desa or dict(Config.DESA_LIST).get(kode_desa,'')

        nid_manual = str(row.get('nasabah_id','')).strip().upper()
        if nid_manual:
            if Nasabah.query.filter_by(nasabah_id=nid_manual).first():
                errors.append(f"Baris {i}: ID {nid_manual} sudah ada"); continue
            nasabah_id = nid_manual
        else:
            nasabah_id = get_next_nasabah_id(kode_desa)

        try:
            import uuid
            n = Nasabah(
                nasabah_id=nasabah_id, jenis='kelompok',
                kode_desa=kode_desa, nama_desa=nama_desa, nama=nama,
                nik=f"NOID-{uuid.uuid4().hex[:12].upper()}",
                no_hp=str(row.get('no_hp','')).strip(),
                alamat=str(row.get('alamat','')).strip(),
                pekerjaan='KELOMPOK', created_by=current_user.id,
            )
            db.session.add(n)
            db.session.flush()
            from ..models import RekeningTabungan as RT
            if not RT.query.filter_by(nasabah_id=n.id).first():
                db.session.add(RT(nasabah_id=n.id, no_rekening=f'TAB-{n.nasabah_id}'))
            ok += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f"Baris {i}: {e}")

    db.session.commit()
    if errors: flash(f'{ok} kelompok berhasil, {len(errors)} error: {"; ".join(errors[:3])}','warning')
    else: flash(f'{ok} nasabah kelompok berhasil diimport.','success')
    return redirect(url_for('import_export.index'))


@import_export_bp.route('/import/anggota', methods=['POST'])
@login_required
def import_anggota():
    """Import anggota kelompok dari Excel."""
    if not current_user.can_write_nasabah(): abort(403)
    from ..models import Nasabah, AnggotaKelompok
    file = request.files.get('file')
    if not file or not file.filename:
        flash('Pilih file terlebih dahulu.','warning')
        return redirect(url_for('import_export.index'))

    rows, err = _read_xlsx_or_csv(file, ['nasabah_id_kelompok','nama','nik','jabatan','no_hp','alamat'])
    if err:
        flash(err,'danger'); return redirect(url_for('import_export.index'))

    ok = 0; errors = []
    for i, row in enumerate(rows,2):
        nid_kel = str(row.get('nasabah_id_kelompok','')).strip().upper()
        nama    = str(row.get('nama','')).strip().upper()
        if not nid_kel or not nama:
            errors.append(f"Baris {i}: nasabah_id_kelompok/nama kosong"); continue
        kelompok = Nasabah.query.filter_by(nasabah_id=nid_kel, jenis='kelompok').first()
        if not kelompok:
            errors.append(f"Baris {i}: kelompok {nid_kel} tidak ditemukan"); continue
        try:
            urut = AnggotaKelompok.query.filter_by(kelompok_id=kelompok.id).count() + 1
            a = AnggotaKelompok(
                kelompok_id=kelompok.id, urut=urut, nama=nama,
                nik=str(row.get('nik','')).strip(),
                jabatan=str(row.get('jabatan','anggota')).strip().lower() or 'anggota',
                no_hp=str(row.get('no_hp','')).strip(),
                alamat=str(row.get('alamat','')).strip(),
            )
            db.session.add(a)
            db.session.flush()
            ok += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f"Baris {i}: {e}")

    db.session.commit()
    if errors: flash(f'{ok} anggota berhasil, {len(errors)} error: {"; ".join(errors[:3])}','warning')
    else: flash(f'{ok} anggota kelompok berhasil diimport.','success')
    return redirect(url_for('import_export.index'))
